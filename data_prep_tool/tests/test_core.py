import tempfile
import unittest
from pathlib import Path

import openpyxl

from app.comparison.workbook_comparator import WorkbookComparator
from app.models.train_type import DataValidationError, ExcelMappingEntry, TrainTypeValues
from app.services.excel_generator import ExcelGenerator
from app.services.train_type_service import TrainTypeService


class FakeRepository:
    def __init__(self, train_type_values: dict | None = None, excel_mapping: dict | None = None) -> None:
        self._train_type_values = train_type_values or {}
        self._excel_mapping = excel_mapping or {}

    def load_train_type_values(self) -> dict:
        return self._train_type_values

    def load_excel_mapping(self) -> dict:
        return self._excel_mapping


class ModelValidationTests(unittest.TestCase):
    def test_train_type_values_from_mapping_rejects_invalid_structure(self) -> None:
        with self.assertRaises(DataValidationError):
            TrainTypeValues.from_mapping("train-a", {1: "invalid"})  # type: ignore[arg-type]

    def test_excel_mapping_entry_validates_cell_reference(self) -> None:
        with self.assertRaises(DataValidationError):
            ExcelMappingEntry.from_mapping({"sheet": "Train Type", "cell": "bad-cell", "field": "speed"})


class ServiceAndGeneratorTests(unittest.TestCase):
    def test_train_type_service_returns_available_types_and_values(self) -> None:
        repository = FakeRepository(
            train_type_values={"train_types": {"generic": {"max_speed_kmh": 160}}},
            excel_mapping={"mappings": []},
        )
        service = TrainTypeService(repository)

        self.assertEqual(service.get_available_train_types(), ["generic"])
        values = service.get_train_type_values("generic")
        self.assertEqual(values.get_value("max_speed_kmh"), 160)

    def test_excel_generator_writes_values_into_template(self) -> None:
        repository = FakeRepository(
            train_type_values={"train_types": {"generic": {"max_speed_kmh": 160}}},
            excel_mapping={
                "mappings": [
                    {"sheet": "Train Type", "cell": "B2", "field": "max_speed_kmh"}
                ]
            },
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            template_path = temp_path / "template.xlsx"
            output_path = temp_path / "output.xlsx"

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Train Type"
            worksheet["B2"] = None
            workbook.save(template_path)

            import app.services.excel_generator as excel_generator_module

            original_template_file = excel_generator_module.TRAIN_TYPE_TEMPLATE_FILE
            excel_generator_module.TRAIN_TYPE_TEMPLATE_FILE = str(template_path)
            try:
                generator = ExcelGenerator(repository)
                train_type_values = TrainTypeValues.from_mapping("generic", {"max_speed_kmh": 160})
                generator.generate(train_type_values, str(output_path))
            finally:
                excel_generator_module.TRAIN_TYPE_TEMPLATE_FILE = original_template_file

            generated_workbook = openpyxl.load_workbook(output_path)
            self.assertEqual(generated_workbook["Train Type"]["B2"].value, 160)

    def test_excel_generator_writes_values_using_parameter_column_mapping(self) -> None:
        repository = FakeRepository(
            train_type_values={"train_types": {"generic": {"train_type_name": "GENERIC_X"}}},
            excel_mapping={
                "mappings": [],
                "parameter_column_mapping": {
                    "sheet": "Train type (2)",
                    "parameter_column": "D",
                    "value_column": "E",
                    "start_row": 1,
                },
            },
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            template_path = temp_path / "template.xlsx"
            output_path = temp_path / "output.xlsx"

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Train type (2)"
            worksheet["D11"] = "train_type_name"
            worksheet["E11"] = "OLD"
            workbook.save(template_path)

            import app.services.excel_generator as excel_generator_module

            original_template_file = excel_generator_module.TRAIN_TYPE_TEMPLATE_FILE
            excel_generator_module.TRAIN_TYPE_TEMPLATE_FILE = str(template_path)
            try:
                generator = ExcelGenerator(repository)
                train_type_values = TrainTypeValues.from_mapping("generic", {"train_type_name": "GENERIC_X"})
                generator.generate(train_type_values, str(output_path))
            finally:
                excel_generator_module.TRAIN_TYPE_TEMPLATE_FILE = original_template_file

            generated_workbook = openpyxl.load_workbook(output_path)
            self.assertEqual(generated_workbook["Train type (2)"]["E11"].value, "GENERIC_X")


class WorkbookComparatorTests(unittest.TestCase):
    def test_comparator_detects_cell_changes_and_missing_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            standard_path = temp_path / "standard.xlsm"
            project_path = temp_path / "project.xlsm"

            standard_wb = openpyxl.Workbook()
            standard_ws = standard_wb.active
            standard_ws.title = "Train type (2)"
            standard_ws["D11"] = "train_type_name"
            standard_ws["E11"] = "STANDARD"
            standard_ws["A1"] = "keep"
            standard_extra = standard_wb.create_sheet("Standard Only")
            standard_extra["A1"] = "exists"
            standard_wb.save(standard_path)

            project_wb = openpyxl.Workbook()
            project_ws = project_wb.active
            project_ws.title = "Train type (2)"
            project_ws["D11"] = "train_type_name"
            project_ws["E11"] = "PROJECT"
            project_ws["A1"] = "keep"
            project_extra = project_wb.create_sheet("Project Only")
            project_extra["A1"] = "exists"
            project_wb.save(project_path)

            comparator = WorkbookComparator(str(standard_path), str(project_path))
            result = comparator.generate_report(str(temp_path / "Workbook_Deviations.xlsx"))

            self.assertEqual(result.summary.total_sheets_compared, 1)
            self.assertGreaterEqual(result.summary.total_cells_compared, 2)
            self.assertEqual(result.summary.total_deviations_found, 3)
            self.assertEqual(result.output_file, str(temp_path / "Workbook_Deviations.xlsx"))

            report_wb = openpyxl.load_workbook(result.output_file)
            report_ws = report_wb["List of Project Deviations"]
            rows = list(report_ws.iter_rows(min_row=2, values_only=True))
            self.assertIn((1, "Missing Sheet", "Project Only", "-", "Missing", "Exists"), rows)
            self.assertIn((2, "Missing Sheet", "Standard Only", "-", "Exists", "Missing"), rows)
            self.assertIn((3, "Value Changed", "Train type (2)", "E11", "STANDARD", "PROJECT"), rows)


if __name__ == "__main__":
    unittest.main()