"""
Run this script once to create templates/train_type_template.xlsx.

Usage:
    python create_template.py
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), "templates")
OUTPUT_PATH = os.path.join(TEMPLATES_DIR, "train_type_template.xlsx")

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
LABEL_FONT = Font(bold=True, size=10)
LABEL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")

ROWS = [
    ("A1", "Parameter",        "B1", "Value"),
    ("A2", "Max Speed (km/h)", "B2", ""),
    ("A3", "Traction Type",    "B3", ""),
    ("A4", "Axle Load (t)",    "B4", ""),
    ("A5", "Length (m)",       "B5", ""),
    ("A6", "Passenger Capacity","B6", ""),
    ("A7", "Power (kW)",       "B7", ""),
    ("A8", "Braking Distance (m)","B8", ""),
    ("A9", "Bogies Count",     "B9", ""),
    ("A10","Gauge (mm)",       "B10",""),
    ("A11","Notes",            "B11",""),
]


def build_template() -> None:
    os.makedirs(TEMPLATES_DIR, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Train Type"

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40

    for label_cell, label_text, value_cell, value_text in ROWS:
        lc = ws[label_cell]
        lc.value = label_text
        lc.alignment = CENTER

        vc = ws[value_cell]
        vc.value = value_text
        vc.alignment = CENTER

        if label_cell == "A1":
            lc.font = HEADER_FONT
            lc.fill = HEADER_FILL
            vc.font = HEADER_FONT
            vc.fill = HEADER_FILL
        else:
            lc.font = LABEL_FONT
            lc.fill = LABEL_FILL

    wb.save(OUTPUT_PATH)
    print(f"Template created: {OUTPUT_PATH}")


if __name__ == "__main__":
    build_template()
