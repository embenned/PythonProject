import logging
import os
import shutil
from dataclasses import dataclass

import openpyxl

from app.config import TRAIN_TYPE_TEMPLATE_FILE
from app.models.train_type import DataValidationError, ExcelMappingEntry, TrainTypeValues
from app.repositories.base_repository import BaseRepository

logger = logging.getLogger(__name__)


@dataclass(slots=True)
class ParameterColumnMapping:
    sheet: str
    parameter_column: str
    value_column: str
    start_row: int = 1
    end_row: int | None = None

    @classmethod
    def from_mapping(cls, raw: dict) -> "ParameterColumnMapping":
        if not isinstance(raw, dict):
            raise DataValidationError("The 'parameter_column_mapping' value must be an object.")

        sheet = raw.get("sheet", "")
        parameter_column = raw.get("parameter_column", "")
        value_column = raw.get("value_column", "")
        start_row = raw.get("start_row", 1)
        end_row = raw.get("end_row")

        if not isinstance(sheet, str) or not sheet.strip():
            raise DataValidationError("The parameter_column_mapping requires a valid 'sheet'.")
        if not isinstance(parameter_column, str) or not parameter_column.strip():
            raise DataValidationError("The parameter_column_mapping requires a valid 'parameter_column'.")
        if not isinstance(value_column, str) or not value_column.strip():
            raise DataValidationError("The parameter_column_mapping requires a valid 'value_column'.")
        if not isinstance(start_row, int) or start_row < 1:
            raise DataValidationError("The parameter_column_mapping 'start_row' must be an integer >= 1.")
        if end_row is not None and (not isinstance(end_row, int) or end_row < start_row):
            raise DataValidationError("The parameter_column_mapping 'end_row' must be null or >= start_row.")

        return cls(
            sheet=sheet,
            parameter_column=parameter_column,
            value_column=value_column,
            start_row=start_row,
            end_row=end_row,
        )


class ExcelGenerator:

    def __init__(self, repository: BaseRepository) -> None:
        self._repository = repository
        self._mapping: list[ExcelMappingEntry] | None = None
        self._parameter_column_mapping: ParameterColumnMapping | None = None

    def _ensure_mapping_loaded(self) -> None:
        if self._mapping is None:
            raw = self._repository.load_excel_mapping()
            mappings = raw.get("mappings", [])
            if not isinstance(mappings, list):
                raise DataValidationError("The top-level 'mappings' value must be a list.")
            self._mapping = [ExcelMappingEntry.from_mapping(entry) for entry in mappings]

            parameter_column_mapping = raw.get("parameter_column_mapping")
            if parameter_column_mapping is not None:
                self._parameter_column_mapping = ParameterColumnMapping.from_mapping(parameter_column_mapping)

    def _write_with_explicit_mappings(self, wb, train_type_values: TrainTypeValues) -> None:
        for entry in self._mapping:
            if entry.sheet not in wb.sheetnames:
                logger.warning("Sheet '%s' not found in template — skipping.", entry.sheet)
                continue

            ws = wb[entry.sheet]
            value = train_type_values.get_value(entry.field)
            if value is None:
                logger.warning(
                    "No value for field '%s' (train type '%s') — cell %s!%s left unchanged.",
                    entry.field,
                    train_type_values.train_type,
                    entry.sheet,
                    entry.cell,
                )
                continue

            ws[entry.cell] = value
            logger.debug("Wrote %s -> %s!%s", entry.field, entry.sheet, entry.cell)

    def _write_with_parameter_column_mapping(self, wb, train_type_values: TrainTypeValues) -> None:
        mapping = self._parameter_column_mapping
        if mapping is None:
            return
        if mapping.sheet not in wb.sheetnames:
            raise DataValidationError(
                f"Sheet '{mapping.sheet}' from parameter_column_mapping was not found in template workbook."
            )

        ws = wb[mapping.sheet]
        end_row = mapping.end_row or ws.max_row

        parameter_to_row: dict[str, int] = {}
        for row in range(mapping.start_row, end_row + 1):
            parameter_key = ws[f"{mapping.parameter_column}{row}"].value
            if isinstance(parameter_key, str) and parameter_key.strip():
                parameter_to_row[parameter_key.strip()] = row

        missing_fields: list[str] = []
        for field_key, field_value in train_type_values.values.items():
            row = parameter_to_row.get(field_key)
            if row is None:
                missing_fields.append(field_key)
                continue
            ws[f"{mapping.value_column}{row}"] = field_value

        if missing_fields:
            logger.warning(
                "The following fields were not found in parameter column '%s' for sheet '%s': %s",
                mapping.parameter_column,
                mapping.sheet,
                ", ".join(sorted(missing_fields)),
            )

    def generate(self, train_type_values: TrainTypeValues, output_path: str) -> None:
        self._ensure_mapping_loaded()

        if not os.path.exists(TRAIN_TYPE_TEMPLATE_FILE):
            raise FileNotFoundError(
                f"Template not found: {TRAIN_TYPE_TEMPLATE_FILE}"
            )

        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        shutil.copy2(TRAIN_TYPE_TEMPLATE_FILE, output_path)
        logger.info("Copied template to: %s", output_path)

        wb = openpyxl.load_workbook(output_path)

        if self._mapping:
            self._write_with_explicit_mappings(wb, train_type_values)
        if self._parameter_column_mapping is not None:
            self._write_with_parameter_column_mapping(wb, train_type_values)

        wb.save(output_path)
        logger.info("Excel file saved: %s", output_path)
