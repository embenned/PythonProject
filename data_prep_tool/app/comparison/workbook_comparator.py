import io
import logging
import os
import time
from pathlib import Path
from typing import Callable, Iterable
from zipfile import BadZipFile

import openpyxl
from openpyxl.utils.cell import get_column_letter, range_boundaries

from app.comparison.models import (
    ComparisonSummary,
    ComparisonError,
    DeviationRecord,
    ReportGenerationError,
    WorkbookComparisonResult,
    WorkbookLoadError,
)
from app.comparison.report_generator import ReportGenerator

try:
    import msoffcrypto
except Exception:  # pragma: no cover - optional dependency
    msoffcrypto = None

logger = logging.getLogger(__name__)

ProgressCallback = Callable[[str], None]

IGNORED_SHEETS = {
    "Cockpit",
    "Revision history",
    "Verify_Import_XML",
}

METADATA_SHEETS = {
    "Passive Rules Check",
}

PASSWORD = "IDDP_Protection_842"


class WorkbookComparator:
    def __init__(
        self,
        standard_workbook_path: str,
        project_workbook_path: str,
        password: str = PASSWORD,
        report_generator: ReportGenerator | None = None,
        progress_callback: ProgressCallback | None = None,
    ) -> None:
        self._standard_workbook_path = standard_workbook_path
        self._project_workbook_path = project_workbook_path
        self._password = password
        self._report_generator = report_generator or ReportGenerator()
        self._progress_callback = progress_callback
        self._standard_workbook = None
        self._project_workbook = None

    def _progress(self, message: str) -> None:
        logger.info(message)
        if self._progress_callback is not None:
            self._progress_callback(message)

    @staticmethod
    def _normalize_value(value):
        if value is None:
            return None
        if isinstance(value, str):
            normalized = value.replace("\r\n", "\n").replace("\r", "\n").strip()
            return normalized if normalized else None
        return value

    @staticmethod
    def _sheet_bounds(worksheet):
        try:
            dimension = worksheet.calculate_dimension()
        except ValueError:
            return None

        if dimension == "A1:A1":
            first_value = worksheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=1, values_only=True)
            first_row = next(first_value, (None,))
            if not first_row or first_row[0] is None:
                return None

        min_col, min_row, max_col, max_row = range_boundaries(dimension)
        return min_col, min_row, max_col, max_row

    def _load_protected_workbook(self, file_path: str):
        if msoffcrypto is None:
            raise WorkbookLoadError(
                "This workbook appears to be password-protected, but the optional decryption library is not available."
            )

        logger.info("Attempting to decrypt protected workbook: %s", file_path)
        with open(file_path, "rb") as source_file:
            office_file = msoffcrypto.OfficeFile(source_file)
            office_file.load_key(password=self._password)
            decrypted_stream = io.BytesIO()
            office_file.decrypt(decrypted_stream)
            decrypted_stream.seek(0)
            return openpyxl.load_workbook(
                decrypted_stream,
                data_only=True,
                read_only=True,
            )

    def _load_workbook(self, file_path: str):
        if not os.path.exists(file_path):
            raise WorkbookLoadError(f"Workbook not found: {file_path}")

        try:
            return openpyxl.load_workbook(
                filename=file_path,
                data_only=True,
                read_only=True,
            )
        except Exception as exc:
            logger.warning("Standard openpyxl load failed for %s: %s", file_path, exc)
            try:
                return self._load_protected_workbook(file_path)
            except Exception as protected_exc:
                raise WorkbookLoadError(f"Unable to load workbook '{file_path}': {protected_exc}") from protected_exc

    def load_workbooks(self):
        self._progress("Loading standard workbook...")
        self._standard_workbook = self._load_workbook(self._standard_workbook_path)
        self._progress("Loading project workbook...")
        self._project_workbook = self._load_workbook(self._project_workbook_path)
        return self._standard_workbook, self._project_workbook

    def close_workbooks(self) -> None:
        for workbook in (self._standard_workbook, self._project_workbook):
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    logger.debug("Workbook close failed during cleanup.", exc_info=True)
        self._standard_workbook = None
        self._project_workbook = None

    @staticmethod
    def _sheet_names_for_comparison(workbook) -> set[str]:
        return {
            sheet_name
            for sheet_name in workbook.sheetnames
            if sheet_name not in IGNORED_SHEETS and sheet_name not in METADATA_SHEETS
        }

    def compare_cells(self, standard_sheet, project_sheet) -> tuple[list[DeviationRecord], int]:
        standard_bounds = self._sheet_bounds(standard_sheet)
        project_bounds = self._sheet_bounds(project_sheet)

        if standard_bounds is None and project_bounds is None:
            return [], 0

        if standard_bounds is None:
            min_col, min_row, max_col, max_row = project_bounds
        elif project_bounds is None:
            min_col, min_row, max_col, max_row = standard_bounds
        else:
            min_col = min(standard_bounds[0], project_bounds[0])
            min_row = min(standard_bounds[1], project_bounds[1])
            max_col = max(standard_bounds[2], project_bounds[2])
            max_row = max(standard_bounds[3], project_bounds[3])

        deviations: list[DeviationRecord] = []
        cells_compared = 0

        standard_rows = standard_sheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )
        project_rows = project_sheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )

        for row_index, (standard_row, project_row) in enumerate(zip(standard_rows, project_rows), start=min_row):
            for col_index, (standard_value, project_value) in enumerate(zip(standard_row, project_row), start=min_col):
                cells_compared += 1
                normalized_standard = self._normalize_value(standard_value)
                normalized_project = self._normalize_value(project_value)
                if normalized_standard != normalized_project:
                    deviations.append(
                        DeviationRecord(
                            change_type="Value Changed",
                            sheet_name=standard_sheet.title,
                            cell_address=f"{get_column_letter(col_index)}{row_index}",
                            standard_value=normalized_standard,
                            project_value=normalized_project,
                        )
                    )

        return deviations, cells_compared

    def compare_sheets(self) -> WorkbookComparisonResult:
        if self._standard_workbook is None or self._project_workbook is None:
            self.load_workbooks()

        standard_sheets = self._sheet_names_for_comparison(self._standard_workbook)
        project_sheets = self._sheet_names_for_comparison(self._project_workbook)

        common_sheets = sorted(standard_sheets & project_sheets)
        standard_only_sheets = sorted(standard_sheets - project_sheets)
        project_only_sheets = sorted(project_sheets - standard_sheets)

        deviations: list[DeviationRecord] = []
        total_cells_compared = 0

        for index, sheet_name in enumerate(common_sheets, start=1):
            self._progress(f"Comparing sheet {index}/{len(common_sheets)}: {sheet_name}")
            sheet_deviations, cells_compared = self.compare_cells(
                self._standard_workbook[sheet_name],
                self._project_workbook[sheet_name],
            )
            deviations.extend(sheet_deviations)
            total_cells_compared += cells_compared

        for sheet_name in standard_only_sheets:
            deviations.append(
                DeviationRecord(
                    change_type="Missing Sheet",
                    sheet_name=sheet_name,
                    cell_address="-",
                    standard_value="Exists",
                    project_value="Missing",
                )
            )

        for sheet_name in project_only_sheets:
            deviations.append(
                DeviationRecord(
                    change_type="Missing Sheet",
                    sheet_name=sheet_name,
                    cell_address="-",
                    standard_value="Missing",
                    project_value="Exists",
                )
            )

        summary = ComparisonSummary(
            total_sheets_compared=len(common_sheets),
            total_cells_compared=total_cells_compared,
            total_deviations_found=len(deviations),
        )
        return WorkbookComparisonResult(deviations=deviations, summary=summary)

    def generate_report(self, output_path: str | None = None) -> WorkbookComparisonResult:
        start_time = time.perf_counter()
        try:
            self._progress("Loading workbooks...")
            self.load_workbooks()
            self._progress("Comparing sheets...")
            result = self.compare_sheets()

            if output_path is None:
                output_path = str(Path(self._project_workbook_path).with_name("Workbook_Deviations.xlsx"))

            self._progress("Creating deviation report...")
            result.output_file = self._report_generator.generate(result, output_path)
            result.summary.execution_time_seconds = round(time.perf_counter() - start_time, 2)
            self._progress(f"Comparison completed in {result.summary.execution_time_seconds:.2f} seconds.")
            return result
        finally:
            self.close_workbooks()

    def save_report(self, result: WorkbookComparisonResult, output_path: str) -> str:
        return self._report_generator.generate(result, output_path)
