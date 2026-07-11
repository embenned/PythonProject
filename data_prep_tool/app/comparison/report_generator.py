import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.comparison.models import ReportGenerationError, WorkbookComparisonResult

logger = logging.getLogger(__name__)


class ReportGenerator:
    REPORT_SHEET_NAME = "List of Project Deviations"

    def generate(self, result: WorkbookComparisonResult, output_path: str) -> str:
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = self.REPORT_SHEET_NAME

            headers = [
                "Change No",
                "Change Type",
                "Sheet Name",
                "Cell Address",
                "Standard Value",
                "Project Value",
            ]
            worksheet.append(headers)

            header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for change_no, deviation in enumerate(
                sorted(result.deviations, key=lambda item: (item.sheet_name or "", item.cell_address or "")),
                start=1,
            ):
                worksheet.append(
                    [
                        change_no,
                        deviation.change_type,
                        deviation.sheet_name,
                        deviation.cell_address,
                        deviation.standard_value,
                        deviation.project_value,
                    ]
                )

            widths = [12, 18, 28, 16, 24, 24]
            for index, width in enumerate(widths, start=1):
                worksheet.column_dimensions[get_column_letter(index)].width = width

            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(output_file)
            logger.info("Deviation report saved to: %s", output_file)
            return str(output_file)
        except Exception as exc:
            logger.exception("Failed to generate deviation report.")
            raise ReportGenerationError(f"Failed to generate deviation report: {exc}") from exc
